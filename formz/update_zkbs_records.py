from zipfile import BadZipFile

from openpyxl import load_workbook

from formz.models import ZkbsOncogene, ZkbsPlasmid


def update_zkbs_celllines(excel_file):
    """
    Takes a ByteIO object for a relevant Excel file from
    the ZKBS and uses its content to update the ZKBS Cell
    Line model
    """

    from zipfile import BadZipFile

    from openpyxl import load_workbook

    from formz.models import ZkbsCellLine

    error_messages = []

    try:
        # Load workbook
        wb = load_workbook(filename=excel_file)

        # Load firt sheet
        sheet = wb.worksheets[0]

        rows = iter(sheet)

        # Skip first row
        next(rows)

        # Get table header
        header = next(rows)
        header_values = [
            str(cell.value).strip().lower() for cell in header if cell.value
        ]

        # Check that row headings are named as expected
        if header_values == [
            "name",
            "synonym",
            "risikogruppe",
            "spezies/organismus",
            "gewebe",
            "virus",
            "gentechnisch verändert",
        ]:
            for row in rows:
                # If cell line not in database, add it
                qs = ZkbsCellLine.objects.filter(name__exact=row[0].value.strip())

                if not qs:
                    ZkbsCellLine.objects.create(
                        name=row[0].value.strip(),
                        synonym=row[1].value if row[1].value else "",
                        organism=row[3].value if row[3].value else "",
                        risk_potential=row[2].value if row[2].value else "",
                        origin=row[4].value if row[4].value else "",
                        virus=row[5].value if row[5].value else "",
                        genetically_modified=True if row[6].value == "Ja" else False,
                    )

                else:
                    if len(qs) == 1:
                        zkbs_cell_line = qs.first()
                        zkbs_cell_line.synonym = row[1].value if row[1].value else ""
                        zkbs_cell_line.organism = row[3].value if row[3].value else ""
                        zkbs_cell_line.risk_potential = (
                            row[2].value if row[2].value else ""
                        )
                        zkbs_cell_line.origin = row[4].value if row[4].value else ""
                        zkbs_cell_line.virus = row[5].value if row[5].value else ""
                        zkbs_cell_line.genetically_modified = (
                            True if row[6].value == "Ja" else False
                        )
                        zkbs_cell_line.save()

        else:
            error_messages.append(
                "The ZKBS cell lines were not updated because the column titles, "
                "in the third row of the file you uploaded, do not match the expected "
                "values: Name, Synonym, Risikogruppe, Spezies/Organismus, Gewebe, "
                "Virus, gentechnisch verändert."
            )

    except (KeyError, BadZipFile):
        error_messages.append(
            "The ZKBS oncogenes were not updated because the file you uploaded "
            "is not an Excel file."
        )

    except Exception as e:
        error_messages.append(e)

    return error_messages


def update_zkbs_plasmids(excel_file):
    """
    Takes a ByteIO object for a relevant Excel file from the ZKBS and
    uses its content to update the ZKBS Plasmid model
    """

    error_messages = []

    try:
        # Load workbook
        wb = load_workbook(filename=excel_file)

        # Load firt sheet
        sheet = wb.worksheets[0]

        rows = iter(sheet)

        # Skip first row
        next(rows)

        # Get table header
        header = next(rows)
        header_values = [
            str(cell.value).strip().lower() for cell in header if cell.value
        ]

        # Check that row headings are named as expected
        if header_values == [
            "name",
            "funktion",
            "herkunft",
            "az zkbs",
            "kurzbeschreibung",
        ]:
            for row in rows:
                # If plasmid not in database, add it, if not update it
                qs = ZkbsPlasmid.objects.filter(name__exact=row[0].value.strip())

                if not qs:
                    ZkbsPlasmid.objects.create(
                        name=row[0].value.strip(),
                        source=row[2].value if row[2].value else "",
                        purpose=row[1].value if row[1].value else "",
                        description=row[4].value if row[4].value else "",
                    )

                else:
                    if len(qs) == 1:
                        zkbs_plasmid = qs.first()
                        zkbs_plasmid.source = row[2].value if row[2].value else ""
                        zkbs_plasmid.purpose = row[1].value if row[1].value else ""
                        zkbs_plasmid.description = row[4].value if row[4].value else ""
                        zkbs_plasmid.save()

        else:
            error_messages.append(
                "The ZKBS plasmids were not updated because the column titles, "
                "in the third row of the file you uploaded, do not match the expected "
                "values: Name, Funktion, Herkunft, AZ ZKBS, Kurzbeschreibung."
            )

    except (KeyError, BadZipFile):
        error_messages.append(
            "The ZKBS oncogenes were not updated because the file you uploaded "
            "is not an Excel file."
        )

    except Exception as e:
        error_messages.append(e)

    return error_messages


def update_zkbs_oncogenes(excel_file):
    """
    Takes a ByteIO object for a relevant Excel file from the ZKBS and
    uses its content to update the ZKBS Oncogenes model
    """

    error_messages = []

    try:
        # Load workbook
        wb = load_workbook(filename=excel_file)

        # Load firt sheet
        sheet = wb.worksheets[0]

        rows = iter(sheet)

        # Skip first row
        next(rows)

        # Get table header
        header = next(rows)
        header_values = [
            str(cell.value).strip().lower() for cell in header if cell.value
        ]

        # Check that row headings are named as expected
        if header_values == [
            "eintragdatum",
            "gen/nukleinsäure",
            "synonym",
            "spezies",
            "bewertung",
            "literatur",
            "zusätzliche maßnahmen",
        ]:
            for row in rows:
                # If oncogene not in database, add it
                qs = ZkbsOncogene.objects.filter(name__exact=row[1].value.strip())

                if not qs:
                    ZkbsOncogene.objects.create(
                        name=row[1]
                        .value.replace("*", "")
                        .strip(),  # next time remove asterisk from all names!!!
                        synonym=row[2].value if row[2].value else "",
                        species=row[3].value if row[3].value else "",
                        risk_potential=row[4].value if row[4].value else "",
                        reference=row[5].value if row[5].value else "",
                        additional_measures=True if row[6].value == "Ja" else False,
                    )

                else:
                    if len(qs) == 1:
                        zkbs_oncogene = qs.first()
                        zkbs_oncogene.synonym = row[2].value if row[2].value else ""
                        zkbs_oncogene.species = row[3].value if row[3].value else ""
                        zkbs_oncogene.risk_potential = (
                            row[4].value if row[4].value else ""
                        )
                        zkbs_oncogene.reference = row[5].value if row[5].value else ""
                        zkbs_oncogene.additional_measures = (
                            True if row[6].value == "Ja" else False
                        )
                        zkbs_oncogene.save()

        else:
            error_messages.append(
                "The ZKBS oncogenes were not updated because the column titles, "
                "in the third row of the file you uploaded, do not match the expected "
                "values: Eintragdatum, Gen/Nukleinsäure, Synonym, Spezies, Bewertung, "
                "Literatur, zusätzliche Maßnahmen."
            )

    except (KeyError, BadZipFile):
        error_messages.append(
            "The ZKBS oncogenes were not updated because the file you uploaded "
            "is not an Excel file."
        )

    except Exception as e:
        error_messages.append(e)

    return error_messages
